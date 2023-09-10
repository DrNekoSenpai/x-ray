from openpyxl import load_workbook
from datetime import datetime
import calendar
from argparse import ArgumentParser

parser = ArgumentParser()
parser.add_argument('-i', '--input', help='Command line input', default=None)
parser.add_argument('-n', '--num_bonuses', help='Number of bonuses available', default=0, type=int)
args = parser.parse_args()

cwl_filename = 'Reddit X-ray CWL Random Distribution.xlsx'
wb = load_workbook(filename=cwl_filename)
sheet_name = wb.sheetnames[0]
sheets = wb[sheet_name]

eligible = []
already_received = []
ineligible = []

for ind, row in enumerate(sheets.iter_rows(values_only=True)):
    if ind == 0:
        headers = row
    else:
        if row[0]:
            eligible.append(row[0])
        if row[1]:
            already_received.append(row[1])
        if row[2]:
            ineligible.append(row[2])

this_month = calendar.month_name[datetime.today().month]
this_year = datetime.today().year

if not eligible and not ineligible:
    # If args.input is provided, use that to input data
    if args.input:
        with open(args.input, 'r') as f:
            errors = []
            try: 
                lines = f.readlines()
                for line in lines:
                    line = line.strip()
                    if line == '':
                        continue
                    player = line.rsplit(' ', 1)
                    if player[0].lower() in [i.lower() for i in already_received]:
                        continue
                    try:
                        player[1] = int(player[1])
                    except:
                        print('Invalid input for number of hits used.')
                        continue

                    missed_hits = player[1]
                    if missed_hits > 2:
                        ineligible.append(player[0])
                    else:
                        eligible.append(player[0])

            except Exception as e: 
                errors.append(e)

        if len(errors) != 0: 
            print('Errors encountered: ')
            for error in errors: 
                print(error)
            raise Exception('Errors encountered. Please check input file.')

    else: 
        print('Enter in names of participating players, with number of hits missed: ')
        for i in range(50):
            player = input(f"[{i+1}] ")
            if player == '':
                break
            player = player.rsplit(' ', 1)

            if player[0].lower() in [i.lower() for i in already_received]:
                continue
            try:
                player[1] = int(player[1])
            except:
                print('Invalid input for number of hits used.')
                continue

            missed_hits = player[1]
            if missed_hits > 2:
                ineligible.append(player[0])
            else:
                eligible.append(player[0])

    print(eligible)
    print(already_received)
    print(ineligible)

    from openpyxl import Workbook
    wb = Workbook()
    ws1 = wb.active
    ws1.title = sheet_name

    ws1.append(headers)
    for ind, val in enumerate(eligible):
        ws1[f"A{ind+2}"] = val
    for ind, val in enumerate(already_received):
        ws1[f"B{ind+2}"] = val
    for ind, val in enumerate(ineligible):
        ws1[f"C{ind+2}"] = val

    wb.save(filename=cwl_filename)

else:
    num_bonuses = int(input("How many bonuses are available this month? "))
    if num_bonuses <= 0: 
        raise ValueError("Number of bonuses must be greater than 0.")

    print(f'**---- {this_month} {this_year} CWL Pseudo-random Distribution ----**')
    print(f'({num_bonuses} available bonuses, total)\n')
    
    print('**Eligible**: ', end='')
    print(', '.join(eligible))
    
    print('**Ineligible (already received)**: ', end='')
    print(', '.join(already_received))
    
    print('**Ineligible (missed hits)**: ', end='')
    print(', '.join(ineligible))