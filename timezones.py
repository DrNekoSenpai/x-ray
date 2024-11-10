from openpyxl import load_workbook
import pandas as pd

# Load the workbook and select the active sheet
file_path = './timezones.xlsx'
wb = load_workbook(filename=file_path, data_only=True)
sheet = wb.active

# Define overlap thresholds
overlap_thresholds = {
    "high": 5, 
    "medium": 4, 
    "low": 3
}

# Initialize the results dictionary for storing overlap levels
overlap_results = {}

# Extract member counts for each time zone (assuming this is in the first column next to each row)
member_counts = {}
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=2):
    timezone = row[1].value  # Column with timezone name (e.g., PST, EST)
    member_count = row[0].value  # Column with the number of members in that timezone
    member_counts[timezone] = member_count

# Iterate over each time block column starting from the 3rd column (time blocks start here)
for col in sheet.iter_cols(min_col=3, max_col=sheet.max_column, min_row=2, max_row=sheet.max_row):
    # Get the time block label from the header row (first row)
    time_block = sheet.cell(row=1, column=col[0].column).value
    
    # Calculate total availability count for this time block
    available_count = sum(member_counts[sheet.cell(row=cell.row, column=2).value]
                          for cell in col if cell.value == "Available")
    
    # Determine overlap level
    if available_count >= overlap_thresholds["high"]:
        overlap = "High Overlap"
    elif available_count >= overlap_thresholds["medium"]:
        overlap = "Medium Overlap"
    elif available_count >= overlap_thresholds["low"]:
        overlap = "Low Overlap"
    else:
        overlap = "No Overlap"
    
    # Store the result for this time block
    overlap_results[time_block] = overlap

# Convert the results to a DataFrame for easy viewing
overlap_df = pd.DataFrame(list(overlap_results.items()), columns=["Time Block", "Overlap Level"])

# Display the results
print(overlap_df)
