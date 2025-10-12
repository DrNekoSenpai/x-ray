import os, json, random, argparse, pandas as pd
from datetime import datetime, timedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Constants
WINNERS_LOG_FILE = "./outputs/winners_log.json"
INPUT_FILE = "./inputs/cwl-input.txt"
BASE_PENALTY_WEEKS = 27
WEEKS_PER_MONTH = 4.5  # Approximate weeks in a month

def read_players(filename):
    """
    Reads player names and hits from a file.
    Expected file format: one player per line as:
       player_name hits
    Returns a list of tuples: (player_name, int(hits))
    """
    players = []
    if not os.path.exists(filename):
        print(f"Error: The file '{filename}' was not found.")
        exit(1)

    with open(filename, 'r', encoding='utf-8') as file:
        for line in file:
            line = line.strip()
            if not line:
                continue
            try:
                name, hits_str = line.rsplit(' ', 1)
                
                name = name.strip()
                hits_done, hits_total = hits_str.split('/')

                players.append((name, int(hits_total), int(hits_done)))

            except ValueError:
                print(f"Skipping invalid formatted line: {line}")

    return players

def build_entries(players, bypass:bool=False):
    """
    Builds a weighted list of entries based on the number of hits.
    Only players that are eligible (or not logged yet) are considered.
    Each player appears in the list as many times as their hits.
    """
    
    # Open the war_bases.xlsx file and check column B for the number of points.
    # Deduct the number of points from the total number of hits (entries).
    war_bases_file = "./inputs/war_bases.xlsx"
    if not os.path.exists(war_bases_file):
        print(f"Error: The file '{war_bases_file}' was not found.")
        exit(1)
    
    month_year = datetime.today().strftime("%Y-%m")
    with open(f"./strikes/inputs/cwl-{month_year}.txt", 'w', encoding='utf-8') as f:
        f.write("")

    try:
        war_bases_data = pd.read_excel(war_bases_file, usecols=["Player Name", "Points"], engine="openpyxl")
        war_bases_dict = dict(zip(war_bases_data["Player Name"], war_bases_data["Points"]))

        # Deduct points from players' hits based on their individual entries in war_bases.xlsx
        for i, (player, hits_total, hits_done) in enumerate(players):
            player_points = war_bases_dict.get(player, 0)  # Default to 0 if player not found
            if bypass: player_points = 0

            if hits_done == 0: adjusted_hits = -1
            else: 
                hits = 5 - hits_total + hits_done
                adjusted_hits = hits - player_points

            # Check if hits done is less than 3. If so, player entries are capped at number of hits done. 
            if hits_done < 3 and adjusted_hits > hits_done:
                adjusted_hits = hits_done

            players[i] = (player, int(adjusted_hits))

            # If player has zero or hits, write in "./strikes/cwl_{month}.txt" and then remove them from the list. 
            if adjusted_hits <= 0:
                month_year = datetime.today().strftime("%Y-%m")
                strikes_file = f"./strikes/inputs/cwl-{month_year}.txt"
                date = datetime.today().strftime("%Y-%m-%d")
                with open(strikes_file, 'a', encoding='utf-8') as strikes_f:
                    strikes_f.write(f"3\n{player}\ny\n7\n{date}\n")
                players[i] = (player, 0)

    except Exception as e:
        print(f"Error processing war bases file: {e}")
        exit(1)
                      
    entries = []
    for player, hits in players:
        # Skip players with zero hits
        if hits <= 0: continue
        if is_eligible(player): entries.extend([player] * hits)

    return entries

def load_winners_log():
    """Loads the winners log from file."""
    if os.path.exists(WINNERS_LOG_FILE):
        with open(WINNERS_LOG_FILE, 'r', encoding='utf-8') as f:
            try:
                return json.load(f)
            except json.JSONDecodeError:
                print("Error: Winners log file is corrupt. Starting with an empty log.")
                return {}
    else:
        return {}

def save_winners_log(log):
    """Saves the winners log to file."""
    with open(WINNERS_LOG_FILE, 'w', encoding='utf-8') as f:
        json.dump(log, f, indent=4)

def weeks_elapsed(from_date_str):
    """Computes weeks elapsed since 'from_date' (YYYY-MM-DD) until today."""
    win_date = datetime.strptime(from_date_str, "%Y-%m-%d")
    now = datetime.today()
    elapsed_days = (now - win_date).days
    return elapsed_days / 7.0

def total_bonus_for_player(log_record):
    """Calculates the total bonus weeks for a player's log record."""
    bonus_entries = log_record.get("bonus", [])
    return sum(entry.get("bonus_weeks", 0) for entry in bonus_entries)

def remaining_weeks_for_record(record):
    elapsed = weeks_elapsed(record["win_date"])
    bonus = total_bonus_for_player(record)
    return BASE_PENALTY_WEEKS - elapsed - bonus

def is_eligible(player_name):
    """
    A player is eligible if they are not in the winners log, or their penalty has elapsed.
    If the penalty has elapsed, remove them from the log immediately.
    """
    winners_log = load_winners_log()
    record = winners_log.get(player_name)
    if not record:
        return True

    remaining = remaining_weeks_for_record(record)
    if remaining <= 0:
        # Purge this single player and persist
        del winners_log[player_name]
        save_winners_log(winners_log)
        return True

    return False

def draw_winners(entries, available_distributions):
    """
    Draws winners randomly based on weighted entries.
    Each player's chance is weighted by the number of hits.
    Once a player wins, all their entries are removed from the pool.
    """
    winners = []
    current_entries = entries.copy()
    
    while available_distributions > 0 and current_entries:
        winner = random.choice(current_entries)
        winners.append(winner)
        # Remove all occurrences of the winning player
        current_entries = [entry for entry in current_entries if entry != winner]
        available_distributions -= 1
    return winners

def record_new_winners(new_winners):
    """
    Records new winners in the winners log with today's date.
    Does not overwrite existing winners.
    """
    winners_log = load_winners_log()
    today_str = datetime.today().strftime("%Y-%m-%d")
    for winner in new_winners:
        if winner not in winners_log:
            winners_log[winner] = {"win_date": today_str, "bonus": []}

    save_winners_log(winners_log)

def purge_reeligible_players():
    """
    Remove any players from the winners log whose remaining penalty weeks <= 0.
    This ensures the log only contains *currently* ineligible players.
    """
    log = load_winners_log()
    to_delete = []
    for name, record in list(log.items()):
        try:
            if remaining_weeks_for_record(record) <= 0:
                to_delete.append(name)
        except Exception:
            # If a record is malformed, err on the side of cleaning it up
            to_delete.append(name)

    if to_delete:
        for name in to_delete:
            del log[name]
        save_winners_log(log)

def update_bonus():
    """
    Processes bonus records from INPUT_FILE.
    Each line should have the format:
         player_name hits

    This function now clears (wipes) any bonus records in winners_log.json
    for the same month as provided by the bonus file, so that if the command is run multiple times
    in the same month, the bonus for that month is refreshed.

                name, hits_str = line.rsplit(' ', 1)
                name = name.strip()
                hits_done, hits_total = hits_str.split('/')

    For each bonus record not already applied, compute:
        hits = 7 - number of hits missed
        bonus_weeks = (hits // 2) + (1 if hits == 7 else 0)

    and update the player's log.
    """
    if not os.path.exists(INPUT_FILE):
        print(f"No bonus file '{INPUT_FILE}' found. Nothing to update.")
        return

    purge_reeligible_players()
    winners_log = load_winners_log()
    changes_made = False

    # Ask user for current month. If not specified, assume current month.
    month = input("Enter the month (YYYY-MM) for bonus processing (default is current month): ").strip()
    if not month: 
        month = datetime.today().strftime("%Y-%m")
    try:
        datetime.strptime(month, "%Y-%m")
    except ValueError:
        print(f"Invalid month format: {month}. Expected format is YYYY-MM.")
        return

    # Read all bonus records in the bonus file first to know which month(s) we'll be processing.
    bonus_records = []
    with open(INPUT_FILE, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                name, hits_str = line.rsplit(' ', 1)

                hits_used, hits_total = hits_str.split('/')
                name = name.strip()

                bonus_records.append((name, month, hits_used, hits_total))

            except ValueError:
                print(f"Skipping invalid bonus line: {line}")
                continue

    # Create a set of months that are being processed from bonus file.
    months_to_update = {record[1] for record in bonus_records}
    
    # Remove all bonus entries for these months from all players in winners_log.
    for name, record in winners_log.items():
        if "bonus" in record:
            original_count = len(record["bonus"])
            # Filter out bonus entries that are in the set of months to update.
            record["bonus"] = [b for b in record["bonus"] if b.get("month") not in months_to_update]
            if len(record["bonus"]) != original_count:
                print(f"Cleared bonus records for {name} for months: {months_to_update}")
                changes_made = True

    # Now process each bonus line from the bonus file.
    for name, month, hits_used, hits_total in bonus_records:
        if name not in winners_log:
            print(f"Player {name} is not in the winners log; skipping bonus update for them.")
            continue

        # Convert hits_used and hits_total to integers.
        try:
            hits_used = int(hits_used)
            hits_total = int(hits_total)

        except ValueError:
            print(f"Invalid hits format for {name}: '{hits_used}/{hits_total}'. Skipping bonus update.")
            continue

        # Scale bonus weeks based on percentage of hits used. 
        # No hits = 0 weeks, 25% of hits used = 1 week, 50% = 2 weeks, 75% = 3 weeks, 100% = 4 weeks.

        percentage_of_hits = hits_used / hits_total if hits_total > 0 else 0
        if percentage_of_hits < 0.25: bonus_weeks = 0
        elif percentage_of_hits < 0.5: bonus_weeks = 1
        elif percentage_of_hits < 0.75: bonus_weeks = 2
        elif percentage_of_hits < 1.0: bonus_weeks = 3
        else: bonus_weeks = 4

        bonus_entry = {"month": month, "bonus_weeks": bonus_weeks}
        winners_log[name].setdefault("bonus", []).append(bonus_entry)
        print(f"Updated bonus for {name} for month {month}: {bonus_weeks} bonus weeks.")
        changes_made = True

    if changes_made:
        save_winners_log(winners_log)
        print("Winners log updated with bonus records.")
    else:
        print("No bonus records were updated.")

def draw_command(available_distributions:int, bypass:bool):
    """Function to perform the draw winners process."""
    purge_reeligible_players()

    players = read_players(INPUT_FILE)
    entries = build_entries(players, bypass)
    if not entries:
        print("No eligible players available for drawing. Exiting draw process.")
        return
    
    month_year = datetime.today().strftime("%Y-%m")
    ineligible_players = []
    for player, _ in players:
        if not is_eligible(player):
            winners_log = load_winners_log()
            record = winners_log.get(player)
            elapsed = weeks_elapsed(record["win_date"])
            bonus = total_bonus_for_player(record)
            remaining = BASE_PENALTY_WEEKS - elapsed - bonus
            ineligible_players.append((player, elapsed, bonus, remaining))

    track = pity_track(players, ineligible_players)
    # For each player in the pity track, if their value is set to TRUE, their entries must double 
    for i, (player, hits) in enumerate(players):
        if track.get(player):
            print(f"Pity track: Doubling entries for {player} who has {hits} hits.")
            players[i] = (player, hits * 2)

    # Sort by threshold ascending. Those with the least time remaining are listed first.
    ineligible_players.sort(key=lambda x: x[3])

    print("---")
    
    winners = draw_winners(entries, available_distributions)
    month = datetime.today().strftime("%B")
    year = datetime.today().strftime("%Y")
    if winners:
        print(f"**Reddit X-ray {month} {year} Weighted Distribution**\n({available_distributions} available bonuses, total{'; FWA base penalties ignored' if bypass else ''})\n")

        eligible_players = [player for player in players if is_eligible(player[0])]
        eligible_players = sorted(eligible_players, key=lambda x: x[1], reverse=True)

        for num in range(12, 0, -1): 
            if len([player[0] for player in eligible_players if player[1] == num]) > 0:
                print(f"{num} entries: {', '.join([player[0] for player in eligible_players if player[1] == num])}")
            
        print("")

        print(f"Ineligible (all values in weeks): ```")
        longest_player_name = max(len(player[0]) for player in ineligible_players)
        if longest_player_name < 11: longest_player_name = 11

        print(f"╔══{'═'*longest_player_name}╤═════════╤═══════╤═══════════╗")
        print(f"║ Player name{' ' * (longest_player_name - 11)} │ Elapsed │ Bonus │ Remaining ║")

        for player, elapsed, bonus, remaining in ineligible_players:
            print(f"║ {player:<{longest_player_name}} │ {elapsed:<7.2f} │ {bonus:<5} │ {remaining:<9.2f} ║")

        print(f"╚══{'═'*longest_player_name}╧═════════╧═══════╧═══════════╝ ```")

        print("")

        print(f"**This month's {available_distributions} selected winners are:**")
        for winner in winners:
            print(f"- {winner}")

        print("---")
        
        # Check if winners have already been recorded for the current month
        winners_log = load_winners_log()
        current_month = datetime.today().strftime("%Y-%m")
        existing_winners = [name for name, record in winners_log.items() if record["win_date"].startswith(current_month)]

        if existing_winners:
            print(f"Winners for {current_month} have already been recorded: {', '.join(existing_winners)}")
            print("Assuming this is a test run. Exiting without recording winners.")
            return

        confirm = input("Confirm distribution? (y/n): ").strip().lower()
        if confirm == 'y':
            record_new_winners(winners)
            print("Winners recorded successfully.")

    else:
        print("No winners could be selected.")

def pity_track(players, ineligible_players):
    wb = load_workbook("pity_track.xlsx", data_only=True)
    ws = wb.active 

    header = [cell.value for cell in ws[1]]
    current_month = datetime.today().strftime("%B %Y")

    last_col = ws.max_column
    last_month = ws.cell(row=1, column=last_col).value

    if last_month != current_month:
        new_col_index = last_col + 1
        ws.cell(row=1, column=new_col_index, value=current_month)
        print(f"Added new column for {current_month} at {get_column_letter(new_col_index)}1")
        last_col = new_col_index

    # Previous code above...
    # Build lowercase name lookup for sheet rows
    name_to_row = {}
    for row in ws.iter_rows(min_row=2):
        name_cell = row[0]
        if name_cell.value:
            name_to_row[str(name_cell.value).strip().lower()] = name_cell.row

    # Check for missing players 
    for player, *_ in players:
        pname = player.strip().lower()
        if pname not in name_to_row:
            # New player: append at the bottom
            new_row = ws.max_row + 1
            ws.cell(row=new_row, column=1, value=player)
            # Mark all previous months as "Not in clan"
            for col_idx in range(2, last_col):
                ws.cell(row=new_row, column=col_idx, value="Not in clan")
                
            print(f"Added new player row for {player} (auto-filled previous months as 'Not in clan').")

            # Add to lookup for potential future updates
            name_to_row[pname] = new_row

    # Now mark ineligible players
    for player, *_ in ineligible_players:
        pname = player.strip().lower()
        row_idx = name_to_row.get(pname)

        if row_idx:
            ws.cell(row=row_idx, column=last_col, value="Ineligible")
        else:
            print(f"Warning: Player {player} not found in pity_track.xlsx")

    # Convert all datetime objects in header to be date strings; i.e. datetime.datetime(2025, 7, 1, 0, 0) -> "7/1/2025"
    for col_idx, cell_value in enumerate(header, start=1):
        if isinstance(cell_value, datetime):
            ws.cell(row=1, column=col_idx, value=f"{cell_value.month}/{cell_value.day}/{cell_value.year}")

    pity_track = {}
    pity_counts = {}
    for player, *_ in players:
        if player in [p[0] for p in ineligible_players]: continue
        pity_counts[player] = 0
        # Find the player’s row
        for row in ws.iter_rows(min_row=2):
            name = str(row[0].value).strip() if row[0].value else ""
            if name.lower() == player.lower():
                # Work backward from (current_month - 1)
                for col in range(last_col - 1, 1, -1):
                    cell_val = row[col-1].value
                    if cell_val in (None, "", 0):
                        pity_counts[player] += 1
                    else:
                        break
                break

        print(f"{player}: {pity_counts[player]} contiguous blanks")
        pity_track[player] = True if pity_counts[player] >= 3 else False

    # Check if winners have already been recorded for the current month
    winners_log = load_winners_log()
    current_month = datetime.today().strftime("%Y-%m")
    existing_winners = [name for name, record in winners_log.items() if record["win_date"].startswith(current_month)]

    if not existing_winners:
        wb.save("pity_track.xlsx")
        print("Pity sheet updated successfully.")

    return pity_track

def main():
    parser = argparse.ArgumentParser(description="CWL Selection Tool")
    parser.add_argument("--update", "-u", action="store_true", help="Update bonus weeks for players.")
    parser.add_argument("--bypass", "-b", action="store_true", help="Bypass check for FWA bases; useful if clan earned immunity")
    parser.add_argument("--distributions", "-d", type=int, help="Number of distributions available", required=False)
    args = parser.parse_args()

    # Check if we have more than 30 players listed in the input file.
    # If so, forcefully set bypass flag to true. 

    players = read_players(INPUT_FILE)

    if args.update: update_bonus()
    else: 
        if args.distributions is not None:
            num_distributions = args.distributions
        else: 
            num_distributions = input("How many distributions were available? ")
        try: num_distributions = int(num_distributions)
        except: return
        
        draw_command(num_distributions, args.bypass)

if __name__ == "__main__":
    main()
